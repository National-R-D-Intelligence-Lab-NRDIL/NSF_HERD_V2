"""
Load institution classification data into raw_institution_classifications.

Downloads REAL data from authoritative sources:
1. IPEDS Directory (via Urban Institute Education Data Portal API)
   - Carnegie classification (cc_basic_2021)
   - Institutional control (public/private)
   - Medical degree flag
   - UNITID (for crosswalk to DoE data)
2. DoE Eligibility Matrix (FY 2025 Excel) — downloaded from ed.gov
   - HSI designation (Title V eligible, 25%+ Hispanic enrollment)
   - HBCU designation (statutory list)
3. AAU/APLU membership — from dbt seed CSVs
4. EPSCoR states — from dbt seed CSV

Crosswalk strategy:
  - Download IPEDS directory (has UNITID + institution name + state)
  - Match IPEDS to HERD institutions by normalized name + state (fuzzy)
  - Use UNITID from IPEDS match to join DoE HSI/HBCU lists

Data sources:
  - Urban Institute API: educationdata.urban.org/api/v1/
  - DoE Eligibility: ed.gov/media/document/ope-2025-eligibility-matrix-*.xlsx

Usage:
    python scripts/load_classifications.py
"""

import asyncio
import csv
import os
import re
import sys
import time
from pathlib import Path

import asyncpg
import openpyxl
import requests
from rapidfuzz import fuzz, process

# Resolve paths
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
SEEDS_DIR = PROJECT_ROOT / "dbt" / "seeds"
DOE_EXCEL_PATH = SCRIPT_DIR / "doe_eligibility_2025.xlsx"

# Urban Institute Education Data Portal API
IPEDS_DIRECTORY_URL = "https://educationdata.urban.org/api/v1/college-university/ipeds/directory/2022/"

# DoE Eligibility Matrix download URL (FY 2025)
DOE_ELIGIBILITY_URL = "https://www.ed.gov/media/document/ope-2025-eligibility-matrix-62325-110283.xlsx"

# Carnegie Basic 2021 classification codes
# Source: carnegieclassifications.acenet.edu
CARNEGIE_R1_CODES = {15}       # Doctoral Universities: Very High Research Activity
CARNEGIE_R2_CODES = {16}       # Doctoral Universities: High Research Activity
CARNEGIE_D_PU_CODES = {17}     # Doctoral/Professional Universities
CARNEGIE_MASTERS_CODES = {18, 19, 20}  # Master's Colleges (Large, Medium, Small)
CARNEGIE_SPECIAL_MED_CODES = {29, 30}  # Special Focus: Medical/Health


def get_db_params():
    """Load database connection params from .env file if present."""
    env_path = PROJECT_ROOT / ".env"
    if env_path.exists():
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ.setdefault(key.strip(), val.strip())
    return {
        "user": os.getenv("POSTGRES_USER", "herd"),
        "password": os.getenv("POSTGRES_PASSWORD", "herd"),
        "database": os.getenv("POSTGRES_DB", "herd_db"),
        "host": os.getenv("POSTGRES_HOST", "localhost"),
        "port": int(os.getenv("POSTGRES_PORT", "5432")),
    }


def load_seed_csv(filename: str) -> set:
    """Load inst_ids from a seed CSV file."""
    path = SEEDS_DIR / filename
    if not path.exists():
        print(f"  Warning: {path} not found, skipping")
        return set()
    inst_ids = set()
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            inst_ids.add(row["inst_id"].strip())
    return inst_ids


def load_epscor_states() -> set:
    """Load EPSCoR state codes from seed CSV."""
    path = SEEDS_DIR / "seed_epscor_states.csv"
    if not path.exists():
        return set()
    states = set()
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["is_epscor"].strip().lower() == "true":
                states.add(row["state"].strip())
    return states


# ─── IPEDS Download ──────────────────────────────────────────────────────────

def download_ipeds_directory() -> list[dict]:
    """Download IPEDS directory from Urban Institute API.

    Returns all doctoral/research institutions (degree_granting, active).
    Paginates if necessary.
    """
    print("  Downloading from Urban Institute Education Data Portal...")
    print(f"  URL: {IPEDS_DIRECTORY_URL}")
    all_results = []
    url = IPEDS_DIRECTORY_URL
    params = {
        "currently_active_ipeds": 1,
        "degree_granting": 1,
    }

    page = 1
    while url:
        print(f"    Fetching page {page}...", end=" ")
        try:
            resp = requests.get(url, params=params if page == 1 else None, timeout=60)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"\n  ERROR: API request failed: {e}")
            print("  Continuing with data collected so far...")
            break

        data = resp.json()
        results = data.get("results", [])
        all_results.extend(results)
        print(f"got {len(results)} records (total: {len(all_results)})")

        url = data.get("next")
        params = None  # next URL already includes params
        page += 1
        if page > 50:  # safety limit
            break
        time.sleep(0.3)  # rate limit courtesy

    print(f"  Total IPEDS records downloaded: {len(all_results)}")
    return all_results


# ─── DoE Eligibility Matrix ──────────────────────────────────────────────────

def download_doe_eligibility():
    """Download the DoE FY2025 Eligibility Matrix Excel file if not cached."""
    if DOE_EXCEL_PATH.exists():
        print(f"  Using cached DoE Excel: {DOE_EXCEL_PATH}")
        return

    print(f"  Downloading DoE Eligibility Matrix from: {DOE_ELIGIBILITY_URL}")
    resp = requests.get(DOE_ELIGIBILITY_URL, timeout=60)
    resp.raise_for_status()
    DOE_EXCEL_PATH.write_bytes(resp.content)
    print(f"  Saved to: {DOE_EXCEL_PATH} ({len(resp.content) / 1024:.0f} KB)")


def extract_doe_hsi_unitids() -> set[str]:
    """Extract UNITID set from the HSI sheet of the DoE eligibility matrix."""
    wb = openpyxl.load_workbook(DOE_EXCEL_PATH, read_only=True)
    ws = wb["HSI"]
    unitids = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row and len(row) > 2 and row[2]:
            uid = str(row[2]).strip()
            if uid.isdigit():
                unitids.add(uid)
    wb.close()
    print(f"  HSI institutions from DoE: {len(unitids)}")
    return unitids


def extract_doe_hbcu_unitids() -> set[str]:
    """Extract UNITID set from the HBCU sheet of the DoE eligibility matrix."""
    wb = openpyxl.load_workbook(DOE_EXCEL_PATH, read_only=True)
    ws = wb["HBCU"]
    unitids = set()
    # Find the header row first
    header_row = None
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
        if row and any(str(c).strip().lower() == "unitid" for c in row if c):
            header_row = i
            break
    start = (header_row or 4) + 1
    for row in ws.iter_rows(min_row=start, values_only=True):
        if row and len(row) > 2 and row[2]:
            uid = str(row[2]).strip()
            if uid.isdigit():
                unitids.add(uid)
    wb.close()
    print(f"  HBCU institutions from DoE: {len(unitids)}")
    return unitids


# ─── Institution Matching ────────────────────────────────────────────────────

def match_institutions(herd_institutions: list[dict], ipeds_data: list[dict]) -> dict:
    """Match HERD institutions to IPEDS records by name + state.

    Uses rapidfuzz for fuzzy matching. Returns {inst_id: ipeds_record}.
    """
    print("  Building IPEDS lookup by state...")

    # Index IPEDS by state
    ipeds_by_state: dict[str, list[tuple[str, dict]]] = {}
    for rec in ipeds_data:
        state = rec.get("state_abbr", "")
        name = rec.get("inst_name", "")
        if state and name:
            if state not in ipeds_by_state:
                ipeds_by_state[state] = []
            ipeds_by_state[state].append((name, rec))

    matched = {}
    unmatched = []
    match_scores = []

    for inst in herd_institutions:
        inst_id = inst["inst_id"]
        herd_name = inst["name"]
        state = inst["state"]

        if state not in ipeds_by_state:
            unmatched.append((inst_id, herd_name, state))
            continue

        candidates = ipeds_by_state[state]
        candidate_names = [c[0] for c in candidates]

        # Exact match first
        for cname, crec in candidates:
            if cname.lower().strip() == herd_name.lower().strip():
                matched[inst_id] = crec
                match_scores.append(100)
                break
        else:
            # Fuzzy match
            result = process.extractOne(
                herd_name, candidate_names, scorer=fuzz.token_sort_ratio, score_cutoff=70
            )
            if result:
                match_name, score, idx = result
                matched[inst_id] = candidates[idx][1]
                match_scores.append(score)
            else:
                unmatched.append((inst_id, herd_name, state))

    avg_score = sum(match_scores) / len(match_scores) if match_scores else 0
    print(f"  Matched: {len(matched)}/{len(herd_institutions)} "
          f"(avg score: {avg_score:.1f}, unmatched: {len(unmatched)})")

    if unmatched:
        n_show = min(15, len(unmatched))
        print(f"  Unmatched institutions ({n_show} of {len(unmatched)}):")
        for uid, name, st in unmatched[:n_show]:
            print(f"    {uid} | {name} ({st})")

    return matched


# ─── Carnegie Classification ─────────────────────────────────────────────────

def classify_carnegie(code) -> str:
    """Map Carnegie Basic 2021 numeric code to label."""
    if code is None:
        return "Unknown"
    try:
        code = int(code)
    except (ValueError, TypeError):
        return "Unknown"
    if code in CARNEGIE_R1_CODES:
        return "R1"
    elif code in CARNEGIE_R2_CODES:
        return "R2"
    elif code in CARNEGIE_D_PU_CODES:
        return "D/PU"
    elif code in CARNEGIE_SPECIAL_MED_CODES:
        return "Special Focus Medical"
    elif code in CARNEGIE_MASTERS_CODES:
        return "D/PU"
    elif code > 0:
        return "D/PU"
    else:
        return "Unknown"


# ─── Main ────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 70)
    print(" LOAD INSTITUTION CLASSIFICATIONS - from authoritative data sources")
    print("=" * 70)

    db_params = get_db_params()
    conn = await asyncpg.connect(**db_params)

    try:
        # 1. Fetch HERD institutions from database
        print("\n[1/7] Loading HERD institutions from database...")
        herd_rows = await conn.fetch("""
            SELECT DISTINCT ON (inst_id) inst_id, name, state
            FROM raw_institutions
            WHERE year = (SELECT MAX(year) FROM raw_institutions)
            ORDER BY inst_id, year DESC
        """)
        herd_list = [dict(r) for r in herd_rows]
        print(f"  {len(herd_list)} HERD institutions found")

        # 2. Download IPEDS directory
        print("\n[2/7] Downloading IPEDS directory (Urban Institute API)...")
        ipeds_data = download_ipeds_directory()

        # 3. Download DoE Eligibility Matrix
        print("\n[3/7] Loading DoE Eligibility Matrix (HSI + HBCU lists)...")
        download_doe_eligibility()
        hsi_unitids = extract_doe_hsi_unitids()
        hbcu_unitids = extract_doe_hbcu_unitids()

        # 4. Match HERD institutions to IPEDS
        print("\n[4/7] Matching HERD -> IPEDS by name + state...")
        matches = match_institutions(herd_list, ipeds_data)

        # 5. Load membership lists from seed CSVs
        print("\n[5/7] Loading membership seeds (AAU, APLU, EPSCoR)...")
        aau_ids = load_seed_csv("seed_aau_members.csv")
        aplu_ids = load_seed_csv("seed_aplu_members.csv")
        epscor_states = load_epscor_states()
        print(f"  AAU: {len(aau_ids)}, APLU: {len(aplu_ids)}, EPSCoR states: {len(epscor_states)}")

        # 6. Build classification records
        print("\n[6/7] Building classification records...")
        rows = []
        stats = {
            "R1": 0, "R2": 0, "D/PU": 0, "Special Focus Medical": 0, "Unknown": 0,
            "public": 0, "private": 0,
            "med": 0, "hbcu": 0, "hsi": 0, "aau": 0, "aplu": 0, "epscor": 0,
        }

        for inst in herd_list:
            inst_id = inst["inst_id"]
            state = inst["state"]
            ipeds_rec = matches.get(inst_id)

            if ipeds_rec:
                unitid = str(ipeds_rec.get("unitid", ""))

                # Carnegie from IPEDS (prefer 2021, fallback to 2018)
                cc_code = ipeds_rec.get("cc_basic_2021") or ipeds_rec.get("cc_basic_2018")
                carnegie = classify_carnegie(cc_code)

                # Control from IPEDS (1=Public, 2=Private nonprofit, 3=Private for-profit)
                ctrl_code = ipeds_rec.get("inst_control", 0)
                control = "Public" if ctrl_code == 1 else "Private" if ctrl_code in (2, 3) else "Unknown"

                # Medical degree from IPEDS
                has_med = bool(ipeds_rec.get("medical_degree", 0))

                # HBCU from DoE official list (matched via UNITID)
                is_hbcu = unitid in hbcu_unitids

                # HSI from DoE official list (matched via UNITID)
                is_hsi = unitid in hsi_unitids
            else:
                unitid = None
                carnegie = "Unknown"
                control = "Unknown"
                has_med = False
                is_hbcu = False
                is_hsi = False

            is_epscor = state in epscor_states

            # Track stats
            stats[carnegie] = stats.get(carnegie, 0) + 1
            if control == "Public": stats["public"] += 1
            elif control == "Private": stats["private"] += 1
            if has_med: stats["med"] += 1
            if is_hbcu: stats["hbcu"] += 1
            if is_hsi: stats["hsi"] += 1
            if inst_id in aau_ids: stats["aau"] += 1
            if inst_id in aplu_ids: stats["aplu"] += 1
            if is_epscor: stats["epscor"] += 1

            rows.append((
                inst_id, unitid, carnegie, control, has_med,
                inst_id in aau_ids, inst_id in aplu_ids,
                is_hbcu, is_hsi, is_epscor,
            ))

        print(f"\n  Classification summary:")
        print(f"    Carnegie: R1={stats['R1']}, R2={stats['R2']}, "
              f"D/PU={stats['D/PU']}, Special Focus Medical={stats['Special Focus Medical']}, "
              f"Unknown={stats['Unknown']}")
        print(f"    Control: Public={stats['public']}, Private={stats['private']}")
        print(f"    Medical school: {stats['med']}")
        print(f"    HBCU: {stats['hbcu']}")
        print(f"    HSI: {stats['hsi']}")
        print(f"    AAU: {stats['aau']}")
        print(f"    APLU: {stats['aplu']}")
        print(f"    EPSCoR: {stats['epscor']}")

        # 7. Write to database
        print(f"\n[7/7] Writing {len(rows)} records to raw_institution_classifications...")
        await conn.execute("DELETE FROM raw_institution_classifications")
        await conn.executemany("""
            INSERT INTO raw_institution_classifications
                (inst_id, unitid, carnegie_class, control, has_med_school,
                 is_aau, is_aplu, is_hbcu, is_hsi, is_epscor)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
            ON CONFLICT (inst_id) DO UPDATE SET
                unitid = EXCLUDED.unitid,
                carnegie_class = EXCLUDED.carnegie_class,
                control = EXCLUDED.control,
                has_med_school = EXCLUDED.has_med_school,
                is_aau = EXCLUDED.is_aau,
                is_aplu = EXCLUDED.is_aplu,
                is_hbcu = EXCLUDED.is_hbcu,
                is_hsi = EXCLUDED.is_hsi,
                is_epscor = EXCLUDED.is_epscor,
                updated_at = NOW()
        """, rows)

        count = await conn.fetchval("SELECT COUNT(*) FROM raw_institution_classifications")
        print(f"\n  Done. {count} rows written.")

    finally:
        await conn.close()

    print("\n" + "=" * 70)
    print(" COMPLETE. Restart the API to pick up the new classifications.")
    print("=" * 70)


if __name__ == "__main__":
    asyncio.run(main())
